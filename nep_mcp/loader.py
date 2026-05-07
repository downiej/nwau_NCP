"""Load IHPA NEP price-weight tables from the published xlsx into typed rows.

The xlsx layout (NEP 2025-26) places one care stream per sheet, each with a
multi-row header and merged cells. Per-sheet readers below skip the header
and emit dataclasses keyed by classification code.

The loader is intentionally the only module that touches openpyxl. Everything
downstream consumes the in-memory store from store.py.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator

import openpyxl


SHEET_ACUTE = "Admitted Acute"
SHEET_SUBACUTE = "Admitted Subacute"
SHEET_MH_ADMITTED = "Admitted Mental Health"
SHEET_MH_COMMUNITY = "Community Mental Health"
SHEET_NON_ADMITTED = "Non-Admitted"
SHEET_AECC = "Emergency - AECC"
SHEET_UDG = "Emergency - UDG"


@dataclass(frozen=True)
class AcuteRow:
    code: str
    description: str
    same_day_eligible: bool
    bundled_icu: bool
    alos: float | None
    inlier_lower: int | None
    inlier_upper: int | None
    paediatric_adjustment: float
    weight_same_day: float | None
    weight_sso_base: float | None
    weight_sso_per_diem: float | None
    weight_inlier: float | None
    weight_lso_per_diem: float | None


@dataclass(frozen=True)
class SubacuteRow:
    code: str
    care_type: str
    episode_type: str
    description: str
    is_same_day_class: bool
    alos: float | None
    inlier_lower: int | None
    inlier_upper: int | None
    weight_same_day: float | None
    weight_sso_per_diem: float | None
    weight_inlier: float | None
    weight_lso_per_diem: float | None


@dataclass(frozen=True)
class MentalHealthAdmittedRow:
    code: str
    description: str
    avg_phase_length: float | None
    inlier_lower: int | None
    inlier_upper: int | None
    weight_sso_base: float | None
    weight_sso_per_diem: float | None
    weight_inlier: float | None
    weight_lso_per_diem: float | None


@dataclass(frozen=True)
class MentalHealthCommunityRow:
    code: str
    description: str
    weight_with_consumer: float | None
    weight_without_consumer: float | None


@dataclass(frozen=True)
class NonAdmittedRow:
    code: str
    description: str
    weight: float
    paediatric_adjustment: float


@dataclass(frozen=True)
class AECCRow:
    code: str
    description: str
    weight: float


@dataclass(frozen=True)
class UDGRow:
    code: str
    description: str
    weight: float


@dataclass(frozen=True)
class PriceWeightTables:
    acute: dict[str, AcuteRow] = field(default_factory=dict)
    subacute: dict[str, SubacuteRow] = field(default_factory=dict)
    mh_admitted: dict[str, MentalHealthAdmittedRow] = field(default_factory=dict)
    mh_community: dict[str, MentalHealthCommunityRow] = field(default_factory=dict)
    non_admitted: dict[str, NonAdmittedRow] = field(default_factory=dict)
    aecc: dict[str, AECCRow] = field(default_factory=dict)
    udg: dict[str, UDGRow] = field(default_factory=dict)


def _f(v) -> float | None:
    if v is None or v == "" or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _i(v) -> int | None:
    f = _f(v)
    return int(f) if f is not None else None


def _flag_yes(v) -> bool:
    return isinstance(v, str) and v.strip().upper() == "YES"


def _is_data_row(code) -> bool:
    if code is None:
        return False
    s = str(code).strip()
    if not s:
        return False
    return any(ch.isalnum() for ch in s) and len(s) <= 8


def _load_acute(ws) -> dict[str, AcuteRow]:
    out: dict[str, AcuteRow] = {}
    for row in ws.iter_rows(min_row=8, values_only=True):
        code = row[1]
        if not _is_data_row(code):
            continue
        out[str(code).strip()] = AcuteRow(
            code=str(code).strip(),
            description=str(row[2] or "").strip(),
            same_day_eligible=_flag_yes(row[3]),
            bundled_icu=_flag_yes(row[4]),
            alos=_f(row[5]),
            inlier_lower=_i(row[6]),
            inlier_upper=_i(row[7]),
            paediatric_adjustment=_f(row[8]) or 1.0,
            weight_same_day=_f(row[9]),
            weight_sso_base=_f(row[10]),
            weight_sso_per_diem=_f(row[11]),
            weight_inlier=_f(row[12]),
            weight_lso_per_diem=_f(row[13]),
        )
    return out


def _load_subacute(ws) -> dict[str, SubacuteRow]:
    out: dict[str, SubacuteRow] = {}
    for row in ws.iter_rows(min_row=8, values_only=True):
        code = row[1]
        if not _is_data_row(code):
            continue
        out[str(code).strip()] = SubacuteRow(
            code=str(code).strip(),
            care_type=str(row[2] or "").strip(),
            episode_type=str(row[3] or "").strip(),
            description=str(row[4] or "").strip(),
            is_same_day_class=_flag_yes(row[5]),
            alos=_f(row[6]),
            inlier_lower=_i(row[7]),
            inlier_upper=_i(row[8]),
            weight_same_day=_f(row[9]),
            weight_sso_per_diem=_f(row[10]),
            weight_inlier=_f(row[11]),
            weight_lso_per_diem=_f(row[12]),
        )
    return out


def _load_mh_admitted(ws) -> dict[str, MentalHealthAdmittedRow]:
    out: dict[str, MentalHealthAdmittedRow] = {}
    for row in ws.iter_rows(min_row=8, values_only=True):
        code = row[1]
        if not _is_data_row(code):
            continue
        out[str(code).strip()] = MentalHealthAdmittedRow(
            code=str(code).strip(),
            description=str(row[2] or "").strip(),
            avg_phase_length=_f(row[3]),
            inlier_lower=_i(row[4]),
            inlier_upper=_i(row[5]),
            weight_sso_base=_f(row[6]),
            weight_sso_per_diem=_f(row[7]),
            weight_inlier=_f(row[8]),
            weight_lso_per_diem=_f(row[9]),
        )
    return out


def _load_mh_community(ws) -> dict[str, MentalHealthCommunityRow]:
    out: dict[str, MentalHealthCommunityRow] = {}
    for row in ws.iter_rows(min_row=8, values_only=True):
        code = row[1]
        if not _is_data_row(code):
            continue
        out[str(code).strip()] = MentalHealthCommunityRow(
            code=str(code).strip(),
            description=str(row[2] or "").strip(),
            weight_with_consumer=_f(row[3]),
            weight_without_consumer=_f(row[4]),
        )
    return out


def _load_non_admitted(ws) -> dict[str, NonAdmittedRow]:
    out: dict[str, NonAdmittedRow] = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        code = row[1]
        weight = _f(row[3])
        if code is None or weight is None:
            continue
        code_str = str(code).strip()
        if not code_str or not any(ch.isdigit() for ch in code_str):
            continue
        out[code_str] = NonAdmittedRow(
            code=code_str,
            description=str(row[2] or "").strip(),
            weight=weight,
            paediatric_adjustment=_f(row[4]) or 1.0,
        )
    return out


def _load_aecc(ws) -> dict[str, AECCRow]:
    out: dict[str, AECCRow] = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        code = row[1]
        weight = _f(row[3])
        if not _is_data_row(code) or weight is None:
            continue
        out[str(code).strip()] = AECCRow(
            code=str(code).strip(),
            description=str(row[2] or "").strip(),
            weight=weight,
        )
    return out


def _load_udg(ws) -> dict[str, UDGRow]:
    out: dict[str, UDGRow] = {}
    for row in ws.iter_rows(min_row=7, values_only=True):
        code = row[1]
        weight = _f(row[3])
        if not _is_data_row(code) or weight is None:
            continue
        out[str(code).strip()] = UDGRow(
            code=str(code).strip(),
            description=str(row[2] or "").strip(),
            weight=weight,
        )
    return out


def load_price_weights(xlsx_path: Path | str) -> PriceWeightTables:
    """Read all NEP price-weight sheets from xlsx_path into one in-memory bundle."""
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Price weight xlsx not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        return PriceWeightTables(
            acute=_load_acute(wb[SHEET_ACUTE]),
            subacute=_load_subacute(wb[SHEET_SUBACUTE]),
            mh_admitted=_load_mh_admitted(wb[SHEET_MH_ADMITTED]),
            mh_community=_load_mh_community(wb[SHEET_MH_COMMUNITY]),
            non_admitted=_load_non_admitted(wb[SHEET_NON_ADMITTED]),
            aecc=_load_aecc(wb[SHEET_AECC]),
            udg=_load_udg(wb[SHEET_UDG]),
        )
    finally:
        wb.close()


def iter_classifications(tables: PriceWeightTables, stream: str) -> Iterator[tuple[str, str]]:
    """Yield (code, description) pairs for every row in the given stream."""
    mapping = {
        "acute": tables.acute,
        "subacute": tables.subacute,
        "mh_admitted": tables.mh_admitted,
        "mh_community": tables.mh_community,
        "non_admitted": tables.non_admitted,
        "aecc": tables.aecc,
        "udg": tables.udg,
    }
    rows = mapping.get(stream)
    if rows is None:
        return
    for code, row in rows.items():
        yield code, getattr(row, "description", "")
